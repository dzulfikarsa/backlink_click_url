import openpyxl
import requests
from bs4 import BeautifulSoup
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

# Fungsi untuk mengambil status URL
def get_url_status(url):
    try:
        response = requests.get(url)
        if response.status_code == 200:
            return '1'  # Mengembalikan '1' untuk status berhasil
        else:
            return 'X'  # Mengembalikan 'X' untuk status gagal
    except requests.exceptions.RequestException:
        return 'X'  # Mengembalikan 'X' untuk status gagal

def process_excel(file_path, url_column):
    chrome_options = Options()
    chrome_options.add_argument("--headless")  # Menjalankan Chrome dalam mode latar belakang
    driver = webdriver.Chrome(options=chrome_options)  # Ganti dengan driver yang sesuai (misalnya Firefox)

    if chrome_options.headless:
        print("Program berjalan dalam mode headless")
    else:
        print("Program berjalan dengan tampilan grafis")

    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active

    header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
    column_index = None
    for index, header in enumerate(header_row):
        if header == url_column:
            column_index = index + 1
            break

    if column_index is None:
        print(f"Tidak dapat menemukan kolom '{url_column}' dalam file Excel.")
        return

    for row in sheet.iter_rows(min_row=2, values_only=True):
        if len(row) >= column_index:
            url = row[column_index - 1]
            click_link(driver, url)
        else:
            print(f"Indeks kolom '{url_column}' di luar jangkauan pada baris: {row}")

    driver.quit()

def click_link(driver, url):
    driver.get(url)

    try:
        elements = driver.find_elements(By.XPATH, f"//a[@href='https://www.budiluhur.ac.id/']")
        if elements:
            for element in elements:
                element.click()
                print(f"Berhasil mengklik tautan di URL: {url}")
        else:
            print(f"Tidak ada tautan dengan URL yang sesuai di URL: {url}")
    except:
        print(f"Gagal menemukan atau mengklik tautan di URL: {url}")

# Membaca file Excel input
input_file_path = 'file/Link V1_24.xlsx'
input_wb = openpyxl.load_workbook(input_file_path)
input_sheet = input_wb.active

# Membuat file Excel output
output_wb = openpyxl.Workbook()
output_sheet = output_wb.active
output_sheet['A1'] = 'URL'
output_sheet['B1'] = 'Status'

# Mengklik URL dan menulis status ke file Excel output
row_num = 2
for row in input_sheet.iter_rows(min_row=2, values_only=True):
    if len(row) >= 1:
        url = row[0]
        output_sheet.cell(row=row_num, column=1).value = url
        status = get_url_status(url)
        output_sheet.cell(row=row_num, column=2).value = status
        row_num += 1

# Menyimpan file Excel output
output_file_path = 'file/Link V1_24_dzul.xlsx'
output_wb.save(output_file_path)


# Ubah file_path ke path file Excel yang sesuai
file_path = output_file_path

# Ubah url_column ke nama kolom yang berisi URL
url_column = 'URL'

# Memanggil fungsi untuk memproses file Excel
process_excel(file_path, url_column)
